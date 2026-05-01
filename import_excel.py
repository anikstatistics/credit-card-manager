"""
One-time import from CreditCard_Dashboard Excel file into SQLite.
Run once:  python import_excel.py
"""
import sys
from pathlib import Path
from datetime import datetime, date

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import db

# Look for the Excel file in the project root first, then the backup folder
_root = Path(__file__).parent
_candidates = [
    _root / "CreditCard_Dashboard_2026-04-21.xlsx",
    _root.parent / "CreditCard_Backups" / "CreditCard_Dashboard_2026-04-21.xlsx",
]
EXCEL_PATH = next((p for p in _candidates if p.exists()), _candidates[0])


# ── Helpers ────────────────────────────────────────────────────────────────────

def _val(cell):
    return cell.value


def _float(cell) -> float:
    v = cell.value
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("৳", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _int(cell) -> int:
    return int(_float(cell))


def _str(cell) -> str:
    v = cell.value
    return str(v).strip() if v is not None else ""


def _date_str(cell) -> str:
    """Return ISO date string YYYY-MM-DD or empty."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.date().isoformat()
        return v.isoformat()
    # Try parsing common text formats
    for fmt in ("%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date().isoformat()
        except ValueError:
            pass
    return str(v).strip()


def _day_from_date_cell(cell) -> int:
    """Extract just the day number from a date cell."""
    v = cell.value
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        return d.day
    s = _date_str(cell)
    if s:
        try:
            return date.fromisoformat(s).day
        except Exception:
            pass
    return 1


# ── Waiver info from Dashboard Waiver Tracker (rows 44-50, by card name) ───────

def _load_waiver_info(ws_dash, ws_cards):
    """Returns {card_id: {waiver_type, waiver_target}} by scanning Dashboard."""
    info = {}

    # Build card_name -> card_id map from Cards sheet
    name_to_id = {}
    for row in ws_cards.iter_rows(min_row=2):
        cid = _str(row[0])
        cname = _str(row[2])
        if cid and cname:
            name_to_id[cname.lower()] = cid

    # Strategy 1: scan entire dashboard for "Waiver_Type" header in ANY column
    wt_col = None
    header_row = None
    for ws_row in ws_dash.iter_rows():
        for cell in ws_row:
            v = str(cell.value or "").strip()
            if v.lower() in ("waiver_type", "waiver type"):
                wt_col = cell.column
                header_row = cell.row
                break
        if header_row:
            break

    if header_row and wt_col:
        # Find Card_Name or Card_ID column in same header row
        name_col = None
        id_col = None
        tgt_col = None
        for cell in ws_dash[header_row]:
            v = str(cell.value or "").strip().lower()
            if v in ("card_name", "card name"):
                name_col = cell.column   # last occurrence wins (picks Waiver Tracker col)
            elif v in ("card_id", "card id"):
                id_col = cell.column
            elif v in ("waiver_target", "waiver target"):
                tgt_col = cell.column

        # Only scan header_row+1 to header_row+20 to avoid drifting into other sections
        max_data_row = header_row + 20
        for ws_row in ws_dash.iter_rows(min_row=header_row + 1, max_row=max_data_row):
            wt_cell = ws_row[wt_col - 1]
            if not wt_cell.value:
                continue
            wt = str(wt_cell.value).strip()
            # Reject values that don't look like valid waiver types (e.g. numbers)
            if wt.lower() in ("waiver_type", "waiver type"):
                continue
            try:
                float(wt)   # if it's a number, skip (efficiency score / wrong col)
                continue
            except ValueError:
                pass

            tgt = _float(ws_row[tgt_col - 1]) if tgt_col else 0.0

            # Get card id
            if id_col:
                cid = _str(ws_row[id_col - 1])
                if cid.startswith("CC"):
                    info[cid] = {"waiver_type": wt, "waiver_target": tgt}
                    continue
            if name_col:
                cname = _str(ws_row[name_col - 1]).lower()
                cid = name_to_id.get(cname)
                if cid:
                    info[cid] = {"waiver_type": wt, "waiver_target": tgt}

    if info:
        return info

    # Strategy 2: hardcoded row scan for the Waiver Tracker section (cols J-P, rows 45-55)
    # This section has: Card_Name | Waiver_Type | Waiver_Target | ...
    # based on the known dashboard layout
    for row_num in range(40, 60):
        try:
            row_vals = [ws_dash.cell(row=row_num, column=c).value for c in range(10, 18)]
        except Exception:
            continue
        # Check if this looks like a data row (first cell = card name, second = waiver type)
        if not row_vals[0]:
            continue
        cname = str(row_vals[0]).strip().lower()
        cid = name_to_id.get(cname)
        if not cid:
            continue
        wt = str(row_vals[1] or "").strip()
        if not wt or wt.lower() in ("waiver_type", "waiver type"):
            continue
        tgt = 0.0
        if row_vals[2] is not None:
            v = row_vals[2]
            try:
                tgt = float(str(v).replace("৳", "").replace(",", "").strip())
            except (ValueError, TypeError):
                tgt = 0.0
        info[cid] = {"waiver_type": wt, "waiver_target": tgt}

    return info


# ── Import functions ───────────────────────────────────────────────────────────

def import_cards(ws_cards, ws_dash):
    print("Importing Cards...")
    waiver_info = _load_waiver_info(ws_dash, ws_cards)
    count = 0
    for row in ws_cards.iter_rows(min_row=2):
        cid = _str(row[0])
        if not cid:
            continue
        bank       = _str(row[1])
        card_name  = _str(row[2])
        card_num   = _str(row[3])
        limit      = _float(row[4])
        issue_mo   = _int(row[5])
        start_bal       = _float(row[6])
        live_outstanding = _float(row[17])   # column R = Live_Outstanding
        stmt_day   = _day_from_date_cell(row[7])
        # util_tgt: may be "30%" string or 0.3 float
        util_raw = row[20].value
        if isinstance(util_raw, str) and "%" in util_raw:
            try:
                util_tgt = float(util_raw.strip().replace("%","")) / 100.0
            except ValueError:
                util_tgt = 0.30
        else:
            util_tgt = _float(row[20])
        points_div = _float(row[23])   # column X = index 23

        wi = waiver_info.get(cid, {})
        wt  = wi.get("waiver_type", "None") or "None"
        wtgt = wi.get("waiver_target", 0.0)

        if util_tgt > 1.0:        # stored as percentage e.g. 30 -> 0.30
            util_tgt /= 100.0

        db.upsert_card({
            "card_id":              cid,
            "bank":                 bank,
            "card_name":            card_name,
            "card_number":          card_num,
            "credit_limit":         limit,
            "issue_month":          issue_mo,
            "starting_balance":     start_bal,
            "balance_sync_amount":  live_outstanding,
            "balance_sync_date":    "",   # set after all txns/payments imported
            "statement_day":        stmt_day,
            "due_days":             15,
            "min_due_percent":      5.0,
            "min_due_fixed":        500.0,
            "points_divisor":       points_div if points_div > 0 else 100.0,
            "utilization_target":   util_tgt if util_tgt > 0 else 0.30,
            "waiver_type":          wt,
            "waiver_target":        wtgt,
            "active":               1,
        })
        count += 1
    print(f"  -> {count} cards imported.")


def import_transactions(ws_txns):
    print("Importing Transactions...")
    count = 0
    for row in ws_txns.iter_rows(min_row=2):
        txn_id  = _str(row[0])
        if not txn_id or not txn_id.startswith("TXN"):
            continue
        dt        = _date_str(row[1])
        cid       = _str(row[2])
        merchant  = _str(row[5])
        inst_id   = _str(row[6])
        txn_type  = _str(row[7])
        category  = _str(row[8])
        amount    = abs(_float(row[9]))   # store as positive
        notes     = _str(row[12])

        if not dt or not cid:
            continue

        db.upsert_transaction({
            "transaction_id":   txn_id,
            "date":             dt,
            "card_id":          cid,
            "merchant":         merchant,
            "installment_id":   inst_id,
            "transaction_type": txn_type if txn_type else "Purchase",
            "category":         category,
            "amount":           round(amount, 2),
            "notes":            notes,
        })
        count += 1
    print(f"  -> {count} transactions imported.")


def import_payments(ws_pays):
    print("Importing Payments...")
    count = 0
    for row in ws_pays.iter_rows(min_row=2):
        pay_id = _str(row[0])
        if not pay_id or not pay_id.startswith("PAY"):
            continue
        dt     = _date_str(row[1])
        cid    = _str(row[2])
        method = _str(row[3])
        amount = abs(_float(row[4]))
        notes  = _str(row[6])

        if not dt or not cid:
            continue

        db.upsert_payment({
            "payment_id":     pay_id,
            "date":           dt,
            "card_id":        cid,
            "payment_method": method,
            "amount":         round(amount, 2),
            "notes":          notes,
        })
        count += 1
    print(f"  -> {count} payments imported.")


def import_installments(ws_insts):
    print("Importing Installments...")
    count = 0
    for row in ws_insts.iter_rows(min_row=2):
        inst_id = _str(row[0])
        if not inst_id or not inst_id.startswith("INST"):
            continue
        cid        = _str(row[1])
        merchant   = _str(row[2])
        start      = _date_str(row[3])
        purchase   = abs(_float(row[4]))
        total_mo   = _int(row[5])
        monthly    = abs(_float(row[6]))
        paid       = _int(row[7])

        if not start or not cid:
            continue

        db.upsert_installment({
            "installment_id":    inst_id,
            "card_id":           cid,
            "merchant":          merchant,
            "start_date":        start,
            "purchase_amount":   round(purchase, 2),
            "total_months":      max(1, total_mo),
            "installment_amount": round(monthly, 2),
            "months_paid":       paid,
        })
        count += 1
    print(f"  -> {count} installments imported.")


def import_rewards(ws_rwds):
    print("Importing Rewards...")
    count = 0
    for row in ws_rwds.iter_rows(min_row=2):
        rw_id = _str(row[0])
        if not rw_id or not rw_id.startswith("RWD"):
            continue
        dt         = _date_str(row[1])
        cid        = _str(row[2])
        redeemed   = abs(_int(row[3]))
        adjustment = _int(row[4])
        notes      = _str(row[6])

        if not cid:
            continue

        db.upsert_reward({
            "reward_id":       rw_id,
            "date":            dt or "2026-01-01",
            "card_id":         cid,
            "points_redeemed": redeemed,
            "adjustment":      adjustment,
            "notes":           notes,
        })
        count += 1
    print(f"  -> {count} rewards imported.")


def import_categories(ws_cats):
    print("Importing Categories...")
    count = 0
    for row in ws_cats.iter_rows(min_row=2):
        kw     = _str(row[0])
        txtype = _str(row[1])
        cat    = _str(row[2])
        ctype  = _str(row[3])
        if not kw or not cat:
            continue
        db.upsert_category({
            "keyword":          kw,
            "transaction_type": txtype or "Purchase",
            "category":         cat,
            "category_type":    ctype or "Spending",
        })
        count += 1
    print(f"  -> {count} categories imported.")


def _set_sync_dates():
    """Set balance_sync_date for each card to day-after its last imported entry."""
    print("Setting balance sync dates...")
    with db.get_conn() as conn:
        cards = conn.execute("SELECT card_id FROM cards").fetchall()
        for row in cards:
            cid = row["card_id"]
            last_txn = conn.execute(
                "SELECT MAX(date) FROM transactions WHERE card_id=?", (cid,)
            ).fetchone()[0]
            last_pay = conn.execute(
                "SELECT MAX(date) FROM payments WHERE card_id=?", (cid,)
            ).fetchone()[0]
            # Pick the later of the two dates
            candidates = [d for d in (last_txn, last_pay) if d]
            if not candidates:
                continue
            last_date = max(candidates)
            # Advance by one day so live_balance starts clean after imports
            from datetime import date as _date, timedelta
            sync_date = (_date.fromisoformat(last_date) + timedelta(days=1)).isoformat()
            conn.execute(
                "UPDATE cards SET balance_sync_date=? WHERE card_id=?",
                (sync_date, cid),
            )
            print(f"  {cid}: sync date = {sync_date}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: File not found: {EXCEL_PATH}")
        print("Please update EXCEL_PATH in import_excel.py to point to your file.")
        sys.exit(1)

    print(f"Opening {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    db.init_db()

    import_cards(wb["Cards"], wb["Dashboard"])
    import_transactions(wb["Transactions"])
    import_payments(wb["Payments"])
    import_installments(wb["Installments"])
    import_rewards(wb["Rewards"])
    import_categories(wb["Categories"])
    _set_sync_dates()

    print("\nImport complete!")
    print(f"Database: {db.DB_PATH}")
    print("\nRun the app with:")
    print("  streamlit run app.py")


if __name__ == "__main__":
    main()
